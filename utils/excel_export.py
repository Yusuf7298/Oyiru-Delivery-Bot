import io
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)

def _header_row(ws, headers: list[str]):
    ws.append(headers)
    ws.row_dimensions[ws.max_row].height = 25
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=12)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

def _fmt_dt(dt) -> str:
    if dt is None:
        return "-"
    if hasattr(dt, "strftime"):
        return dt.strftime("%Y-%m-%d %H:%M")
    return str(dt)

def _fmt_day(dt) -> str:
    if dt is None:
        return "-"
    if hasattr(dt, "strftime"):
        return dt.strftime("%A (%Y-%m-%d)")
    return str(dt)

def generate_excel(orders: list, stats: dict, returns_map: dict = None) -> bytes:
    if returns_map is None:
        returns_map = {}

    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Hotel Delivery & Returns Report
    # ----------------------------------------------------
    ws_hotels = wb.active
    ws_hotels.title = "Hotel Delivery & Returns"
    ws_hotels.freeze_panes = "A2"
    
    _header_row(ws_hotels, [
        "Hotel Name",
        "Ordered Person (Customer)",
        "Customer Phone",
        "Product Delivered",
        "Delivered Quantity (KG)",
        "Delivery Day",
        "Delivery Date & Time",
        "Returned Product / Reason",
        "Order Status"
    ])

    for order in orders:
        hotel_name = order.hotel.name if order.hotel else "N/A"
        customer_name = order.customer.full_name if order.customer else "N/A"
        customer_phone = order.customer.phone if order.customer else "N/A"
        ret_info = returns_map.get(order.id, "-")
        deliv_date = _fmt_dt(order.delivered_at or order.created_at)
        deliv_day = _fmt_day(order.delivered_at or order.created_at)

        if order.items:
            for item in order.items:
                prod_name = item.product.name if item.product else "Uploaded List"
                qty = item.quantity if item.product else "N/A"
                ws_hotels.append([
                    hotel_name, customer_name, customer_phone,
                    prod_name, qty, deliv_day, deliv_date,
                    ret_info, order.status.value
                ])
                r = ws_hotels.max_row
                for cell in ws_hotels[r]:
                    cell.border = THIN_BORDER
                if r % 2 == 0:
                    for cell in ws_hotels[r]:
                        cell.fill = _ALT_FILL
        else:
            fname = getattr(order, "original_filename", None) or "Uploaded File"
            ws_hotels.append([
                hotel_name, customer_name, customer_phone,
                f"File: {fname}", "N/A", deliv_day, deliv_date,
                ret_info, order.status.value
            ])
            r = ws_hotels.max_row
            for cell in ws_hotels[r]:
                cell.border = THIN_BORDER

    _auto_width(ws_hotels)

    # ----------------------------------------------------
    # Sheet 2: Delivery Partners Report
    # ----------------------------------------------------
    ws_drivers = wb.create_sheet("Delivery Partners")
    ws_drivers.freeze_panes = "A2"
    
    _header_row(ws_drivers, [
        "Delivery Partner / Driver",
        "Product Delivered",
        "Delivered Quantity (KG)",
        "Destination Hotel",
        "Recipient Customer",
        "Customer Phone",
        "Delivery Date & Time",
        "Returned Product / Reason",
        "Delivery Status"
    ])

    for order in orders:
        driver_name = order.driver_name or (order.delivery_partner.user.full_name if (order.delivery_partner and getattr(order.delivery_partner, "user", None)) else "Unassigned")
        hotel_name = order.hotel.name if order.hotel else "N/A"
        customer_name = order.customer.full_name if order.customer else "N/A"
        customer_phone = order.customer.phone if order.customer else "N/A"
        ret_info = returns_map.get(order.id, "-")
        deliv_date = _fmt_dt(order.delivered_at or order.created_at)

        if order.items:
            for item in order.items:
                prod_name = item.product.name if item.product else "File Order"
                qty = item.quantity if item.product else "N/A"
                ws_drivers.append([
                    driver_name, prod_name, qty, hotel_name,
                    customer_name, customer_phone, deliv_date,
                    ret_info, order.status.value
                ])
                r = ws_drivers.max_row
                for cell in ws_drivers[r]:
                    cell.border = THIN_BORDER
                if r % 2 == 0:
                    for cell in ws_drivers[r]:
                        cell.fill = _ALT_FILL
        else:
            fname = getattr(order, "original_filename", None) or "Uploaded File"
            ws_drivers.append([
                driver_name, f"File: {fname}", "N/A", hotel_name,
                customer_name, customer_phone, deliv_date,
                ret_info, order.status.value
            ])
            r = ws_drivers.max_row
            for cell in ws_drivers[r]:
                cell.border = THIN_BORDER

    _auto_width(ws_drivers)

    # ----------------------------------------------------
    # Sheet 3: Store Managers Weekly Report
    # ----------------------------------------------------
    ws_sm = wb.create_sheet("Store Managers Weekly")
    ws_sm.freeze_panes = "A2"

    _header_row(ws_sm, [
        "Hotel Name",
        "Total Orders (Weekly)",
        "Successfully Delivered Orders",
        "Total Delivered Weight (KG)",
        "Returned Orders Count",
        "Returned Products & Reasons",
        "Delivery Completion Rate (%)"
    ])

    hotel_stats = {}
    for order in orders:
        h_name = order.hotel.name if order.hotel else "Unassigned"
        if h_name not in hotel_stats:
            hotel_stats[h_name] = {
                "total": 0, "delivered": 0, "kg": 0.0, "returned_cnt": 0, "returns": []
            }
        
        hotel_stats[h_name]["total"] += 1
        
        if order.status.value == "Delivered":
            hotel_stats[h_name]["delivered"] += 1
            if order.items:
                for item in order.items:
                    try:
                        hotel_stats[h_name]["kg"] += float(item.quantity or 0)
                    except ValueError:
                        pass
        
        if order.id in returns_map:
            hotel_stats[h_name]["returned_cnt"] += 1
            hotel_stats[h_name]["returns"].append(returns_map[order.id])

    for h_name, h_data in hotel_stats.items():
        total_cnt = h_data["total"]
        deliv_cnt = h_data["delivered"]
        rate = round((deliv_cnt / total_cnt * 100), 1) if total_cnt > 0 else 0.0
        ret_summary = "; ".join(h_data["returns"]) if h_data["returns"] else "-"
        
        ws_sm.append([
            h_name, total_cnt, deliv_cnt,
            round(h_data["kg"], 2), h_data["returned_cnt"],
            ret_summary, f"{rate}%"
        ])
        r = ws_sm.max_row
        for cell in ws_sm[r]:
            cell.border = THIN_BORDER
        if r % 2 == 0:
            for cell in ws_sm[r]:
                cell.fill = _ALT_FILL

    _auto_width(ws_sm)

    # ----------------------------------------------------
    # Sheet 4: Master Orders Overview
    # ----------------------------------------------------
    ws_master = wb.create_sheet("Master Orders Overview")
    ws_master.freeze_panes = "A2"

    _header_row(ws_master, [
        "Order Number", "Hotel Name", "Customer Name", "Customer Phone",
        "Status", "Driver Name", "Submitted Date", "Delivered Date",
        "Rating", "Feedback", "Returned Products", "Products / Items Summary"
    ])

    for order in orders:
        h_name = order.hotel.name if order.hotel else "N/A"
        c_name = order.customer.full_name if order.customer else "N/A"
        c_phone = order.customer.phone if order.customer else "N/A"
        ret_info = returns_map.get(order.id, "-")
        
        prod_str = ""
        if order.items:
            prod_str = "; ".join(
                f"{item.product.name} ({item.quantity} {getattr(item.product, 'unit', 'KG')})"
                for item in order.items if item.product
            )
        else:
            prod_str = f"File: {getattr(order, 'original_filename', None) or 'Uploaded List'}"

        ws_master.append([
            order.order_number, h_name, c_name, c_phone,
            order.status.value, order.driver_name or "Unassigned",
            _fmt_dt(order.created_at), _fmt_dt(order.delivered_at),
            order.rating or "-", order.feedback or "-",
            ret_info, prod_str
        ])
        r = ws_master.max_row
        for cell in ws_master[r]:
            cell.border = THIN_BORDER
        if r % 2 == 0:
            for cell in ws_master[r]:
                cell.fill = _ALT_FILL

    _auto_width(ws_master)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_customer_excel(customer, orders: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Orders"
    ws.freeze_panes = "A2"

    _header_row(ws, [
        "Order Number",
        "Hotel Name",
        "Products / File",
        "Total Items / Quantity",
        "Order Status",
        "Placed Date",
        "Delivered Date",
        "Driver",
        "Note"
    ])

    for order in orders:
        h_name = order.hotel.name if getattr(order, "hotel", None) else (getattr(customer, "hotel", None).name if getattr(customer, "hotel", None) else "N/A")
        status_val = order.status.value if hasattr(order.status, "value") else str(order.status)
        placed_dt = _fmt_dt(order.created_at)
        deliv_dt = _fmt_dt(order.delivered_at)
        driver_str = order.driver_name or "-"
        note_str = order.note or "-"

        prod_lines = []
        total_kg = 0.0
        if order.items:
            for item in order.items:
                if item.product:
                    prod_lines.append(f"{item.product.name} ({item.quantity} {getattr(item.product, 'unit', 'KG')})")
                    try:
                        total_kg += float(item.quantity)
                    except (ValueError, TypeError):
                        pass
            prod_summary = "; ".join(prod_lines)
            qty_summary = f"{total_kg:.1f} KG" if total_kg > 0 else f"{len(order.items)} items"
        else:
            fname = getattr(order, "original_filename", None) or "Uploaded List"
            ftype = (getattr(order, "file_type", None) or "file").upper()
            prod_summary = f"[{ftype}] {fname}"
            qty_summary = "1 file"

        ws.append([
            order.order_number,
            h_name,
            prod_summary,
            qty_summary,
            status_val,
            placed_dt,
            deliv_dt,
            driver_str,
            note_str
        ])
        r = ws.max_row
        for cell in ws[r]:
            cell.border = THIN_BORDER
        if r % 2 == 0:
            for cell in ws[r]:
                cell.fill = _ALT_FILL

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_driver_excel(driver, orders: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delivery History"
    ws.freeze_panes = "A2"

    _header_row(ws, [
        "Order Number",
        "Destination Hotel",
        "Hotel Address",
        "Customer Name",
        "Customer Phone",
        "Products / File",
        "Delivery Status",
        "Accepted At",
        "Delivered At",
        "Order Note"
    ])

    for order in orders:
        h_name = order.hotel.name if getattr(order, "hotel", None) else "N/A"
        h_addr = order.hotel.address if getattr(order, "hotel", None) else "-"
        c_name = order.customer.full_name if getattr(order, "customer", None) else "N/A"
        c_phone = order.customer.phone if getattr(order, "customer", None) else "-"
        status_val = order.status.value if hasattr(order.status, "value") else str(order.status)
        acc_dt = _fmt_dt(order.accepted_at)
        deliv_dt = _fmt_dt(order.delivered_at)
        note_str = order.note or "-"

        prod_lines = []
        if order.items:
            for item in order.items:
                if item.product:
                    prod_lines.append(f"{item.product.name} ({item.quantity} {getattr(item.product, 'unit', 'KG')})")
            prod_summary = "; ".join(prod_lines)
        else:
            fname = getattr(order, "original_filename", None) or "Uploaded List"
            prod_summary = f"File: {fname}"

        ws.append([
            order.order_number,
            h_name,
            h_addr,
            c_name,
            c_phone,
            prod_summary,
            status_val,
            acc_dt,
            deliv_dt,
            note_str
        ])
        r = ws.max_row
        for cell in ws[r]:
            cell.border = THIN_BORDER
        if r % 2 == 0:
            for cell in ws[r]:
                cell.fill = _ALT_FILL

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_hotel_orders_excel(hotel, orders: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hotel Orders Report"
    ws.freeze_panes = "A2"

    hotel_name = hotel.name if hasattr(hotel, "name") else (str(hotel) if hotel else "Hotel")

    _header_row(ws, [
        "Order Number",
        "Hotel Name",
        "Ordered By (Staff)",
        "Staff Phone",
        "Products Breakdown",
        "Quantity Summary",
        "Order Status",
        "Order Time",
        "Delivery Time",
        "Assigned Driver",
        "Note"
    ])

    for order in orders:
        c_name = order.customer.full_name if getattr(order, "customer", None) else "—"
        c_phone = order.customer.phone if getattr(order, "customer", None) and order.customer.phone else "—"
        status_val = order.status.value if hasattr(order.status, "value") else str(order.status)
        order_dt = _fmt_dt(order.created_at)
        deliv_dt = _fmt_dt(order.delivered_at)
        driver_str = order.driver_name or (order.delivery_partner.full_name if getattr(order, "delivery_partner", None) else "—")
        note_str = order.note or "—"

        prod_lines = []
        total_kg = 0.0
        if order.items:
            for item in order.items:
                if item.product:
                    unit_str = getattr(item.product, "unit", "KG") or "KG"
                    prod_lines.append(f"{item.product.name} ({item.quantity} {unit_str})")
                    try:
                        total_kg += float(item.quantity)
                    except (ValueError, TypeError):
                        pass
            prod_summary = "; ".join(prod_lines)
            qty_summary = f"{total_kg:.1f} KG" if total_kg > 0 else f"{len(order.items)} items"
        else:
            fname = getattr(order, "original_filename", None) or "Uploaded File"
            ftype = (getattr(order, "file_type", None) or "file").upper()
            prod_summary = f"[{ftype}] {fname}"
            qty_summary = "1 file"

        ws.append([
            order.order_number,
            hotel_name,
            c_name,
            c_phone,
            prod_summary,
            qty_summary,
            status_val,
            order_dt,
            deliv_dt,
            driver_str,
            note_str
        ])
        r = ws.max_row
        for cell in ws[r]:
            cell.border = THIN_BORDER
        if r % 2 == 0:
            for cell in ws[r]:
                cell.fill = _ALT_FILL

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


